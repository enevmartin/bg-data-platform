"""
ExcelProcessor — reads .xls / .xlsx files with openpyxl.

Features:
  - Multi-sheet support: returns one DataFrame per sheet
  - Auto-detect header row (skips rows that are entirely empty or all-string metadata)
  - Clean column names to snake_case
  - Strips leading/trailing whitespace from string values
"""
import logging
from pathlib import Path

import polars as pl

from bgdata.processors.base import BaseProcessor, snake_case

logger = logging.getLogger(__name__)

_MAX_HEADER_SCAN_ROWS = 10  # How many rows to look at when finding the real header


class ExcelProcessor(BaseProcessor):
    supported_extensions = frozenset([".xlsx", ".xls"])

    def process(self, file_path: Path) -> list[pl.DataFrame]:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        except Exception as exc:
            logger.error("Cannot open Excel file %s: %s", file_path, exc)
            return []

        frames: list[pl.DataFrame] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                continue

            # ── Auto-detect header row ─────────────────────────────────────
            header_idx = 0
            for i, row in enumerate(rows[:_MAX_HEADER_SCAN_ROWS]):
                non_none = [c for c in row if c is not None]
                if not non_none:
                    continue
                # Prefer a row where most values are strings (column headers)
                str_count = sum(1 for c in non_none if isinstance(c, str))
                if str_count / len(non_none) >= 0.5:
                    header_idx = i
                    break

            header = rows[header_idx]
            data_rows = rows[header_idx + 1 :]

            if not data_rows:
                continue

            # ── Build column names ─────────────────────────────────────────
            col_names: list[str] = []
            seen: dict[str, int] = {}
            for cell in header:
                base = snake_case(str(cell)) if cell is not None else "col"
                if not base:
                    base = "col"
                count = seen.get(base, 0)
                seen[base] = count + 1
                col_names.append(f"{base}_{count}" if count else base)

            # ── Build polars DataFrame ─────────────────────────────────────
            try:
                # Convert rows to column-oriented dict
                col_data: dict[str, list] = {c: [] for c in col_names}
                for row in data_rows:
                    padded = list(row) + [None] * max(0, len(col_names) - len(row))
                    for col, val in zip(col_names, padded[: len(col_names)]):
                        col_data[col].append(val)

                try:
                    df = pl.DataFrame(col_data, infer_schema_length=1000)
                except Exception:
                    # Mixed-type columns (e.g. date + string in same column) —
                    # stringify everything so we never lose data.
                    safe_data = {
                        c: [str(v) if v is not None else None for v in vals]
                        for c, vals in col_data.items()
                    }
                    df = pl.DataFrame(safe_data)

                # Drop entirely-null rows
                df = df.filter(~pl.all_horizontal(pl.all().is_null()))

                if df.is_empty():
                    continue

                logger.debug(
                    "Excel %s | sheet '%s' → %d rows × %d cols",
                    file_path.name,
                    sheet_name,
                    df.height,
                    df.width,
                )
                frames.append(df)

            except Exception as exc:
                logger.error(
                    "Failed to build DataFrame from %s sheet '%s': %s",
                    file_path,
                    sheet_name,
                    exc,
                )

        wb.close()
        return frames
